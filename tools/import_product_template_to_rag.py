"""tools/import_product_template_to_rag.py — імпорт заповненого Excel-шаблону в RAG бота.

Що робить:
1. Читає docs/product_template_emet_правки.xlsx (заповнений медвідділом + Sales Director)
2. Для кожного продукту створює дві структуровані картки .md:
   - {product}_clinical.md — медичні факти (склад/протипоказ/протокол/побічки/догляд)
   - {product}_sales.md — sales playbook (USP/конкуренти/заперечення/cross-sell)
3. Зберігає у data/manual_product_cards/  — постійне сховище
4. Додає чанки в products_openai ChromaDB-індекс з метаданими:
   - product_canonical: канонічна назва
   - scope: 'product'
   - source: '[KARTKA] {product} — {section}'
5. Чанки переживають auto-sync завдяки інтеграції в sync_manager (окремий patch потрібен)

Використання (на сервері, в Docker):
    docker cp docs/product_template_emet_правки.xlsx emet_bot_app:/tmp/template.xlsx
    docker exec emet_bot_app python /app/tools/import_product_template_to_rag.py \\
        --xlsx /tmp/template.xlsx --rebuild

Або локально для preview:
    python tools/import_product_template_to_rag.py --xlsx docs/product_template_emet_правки.xlsx --dry-run
"""
import argparse
import os
import re
import sys
from pathlib import Path
from openpyxl import load_workbook


# Маппінг колонок Excel → ключі для .md
# (key, header_in_excel, section: 'clinical' | 'sales' | 'meta')
COLUMN_MAP = [
    ("canonical", "Канонічна назва", "meta"),
    ("variants_writing", "Варіанти написання (UA+EN)", "meta"),
    ("manufacturer", "Виробник + країна", "clinical"),
    ("category", "Категорія", "meta"),
    ("certification", "Сертифікація (CE / МОЗ / INCI)", "clinical"),
    ("composition", "Активний склад + концентрація", "clinical"),
    ("volume", "Об'єм упаковки", "clinical"),
    ("volume_form", "Форма випуску", "clinical"),
    ("mechanism_of_action", "Механізм дії", "clinical"),
    ("duration_effect", "Тривалість ефекту (місяці)", "clinical"),
    ("onset", "Коли видно результат", "clinical"),
    ("indications", "Показання", "clinical"),
    ("contraindications_absolute", "Абсолютні протипоказання", "clinical"),
    ("contraindications_relative", "Відносні протипоказання", "clinical"),
    ("pregnancy_lactation", "Вагітність / лактація", "clinical"),
    ("age_restrictions", "Вікові обмеження", "clinical"),
    ("zones_allowed", "Дозволені зони (✅)", "clinical"),
    ("zones_forbidden", "ЗАБОРОНЕНІ зони (⛔)", "clinical"),
    ("injection_depth", "Глибина введення", "clinical"),
    ("technique", "Техніка", "clinical"),
    ("dosage", "Дозування за процедуру", "clinical"),
    ("course_count", "Кількість процедур у курсі", "clinical"),
    ("interval_repeat", "Інтервал між процедурами", "clinical"),
    ("compatibility_emet", "Сумісність з EMET-препаратами", "clinical"),
    ("compatibility_devices", "Сумісність з апаратами (HIFU/RF/лазер)", "clinical"),
    ("recovery", "Реабілітація / down-time", "clinical"),
    ("post_procedure_care", "Догляд ПІСЛЯ процедури", "clinical"),
    ("side_effects_common", "Поширені побічні (норма)", "clinical"),
    ("storage_temperature", "Температура зберігання", "clinical"),
    ("price_segment", "Ціновий сегмент (бюджет/середній/преміум)", "sales"),
    ("usp", "USP (1 речення — головна перевага)", "sales"),
    ("vs_competitors", "VS конкуренти (sales-кут)", "sales"),
    ("common_objections", "ТОП-3 заперечення лікарів + аргумент", "sales"),
    ("killer_phrase", "Killer phrase (готова фраза для лікаря)", "sales"),
    ("cross_sell", "З якими EMET-препаратами продається разом", "sales"),
    ("ideal_patient", "Профіль ідеального пацієнта (psycho + клініка)", "sales"),
    ("before_after_url", "Посилання на До/Після (links.emet.in.ua)", "sales"),
]


# Маппінг канонічної назви Excel → product_canonical у RAG
def canonical_for_rag(excel_name: str) -> str:
    """Excel name → product_canonical для metadata.
    Має співпадати з тим що використовує classifier і RAG retrieval."""
    name = (excel_name or "").strip()
    name_low = name.lower()
    if name_low.startswith("ellansé"):
        return "Ellansé"
    if "petaran" in name_low:
        return "Petaran"
    if "neuramis" in name_low:
        return "Neuramis"
    if "vitaran exosome" in name_low or "skin healer" in name_low or "azulene" in name_low or "sleeping" in name_low:
        return "Vitaran Skin Healer"
    if "vitaran whitening" in name_low or "hp cell vitaran whitening" in name_low:
        return "HP Cell Vitaran Whitening"
    if "vitaran tox" in name_low:
        return "HP Cell Vitaran Tox Eye"
    if "hp cell vitaran" in name_low:
        return "HP Cell Vitaran i"
    if "exoxe" in name_low:
        return "EXOXE"
    if "skinbooster" in name_low:
        return "IUSE SKINBOOSTER HA 20"
    if "hair regrowth" in name_low or "iuse hair" in name_low:
        return "IUSE HAIR REGROWTH"
    if "iuse collagen" in name_low:
        return "IUSE Collagen"
    if "neuronox" in name_low:
        return "Neuronox"
    if "magnox" in name_low:
        return "Magnox"
    return name  # fallback


def read_xlsx(xlsx_path: Path) -> list[dict]:
    """Читає Excel і повертає список dict-ів по продуктам."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb["📦 Продукти"]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    # Mapping header → key by COLUMN_MAP
    header_to_key = {h: k for k, h, _ in COLUMN_MAP}
    products = []
    for r in range(2, ws.max_row + 1):
        product = {}
        for c, header in enumerate(headers, 1):
            key = header_to_key.get(header)
            if not key:
                continue
            val = ws.cell(row=r, column=c).value
            product[key] = (str(val).strip() if val is not None else "")
        if product.get("canonical"):
            products.append(product)
    return products


def make_clinical_md(p: dict) -> str:
    """Створює медичну картку продукту як .md з YAML header."""
    canonical = canonical_for_rag(p["canonical"])
    safe_name = p["canonical"][:80]
    yaml_header = (
        f"---\n"
        f"product_name: {safe_name!r}\n"
        f"product_canonical: {canonical!r}\n"
        f"section: clinical\n"
        f"source: med_dept_template_2026\n"
        f"---\n\n"
    )
    body = f"# {safe_name} — Медична картка\n\n"
    if p.get("manufacturer"):
        body += f"**Виробник:** {p['manufacturer']}\n"
    if p.get("certification"):
        body += f"**Сертифікація:** {p['certification']}\n"
    if p.get("category"):
        body += f"**Категорія:** {p['category']}\n"
    body += "\n## Склад і форма\n"
    if p.get("composition"):
        body += f"- **Активний склад:** {p['composition']}\n"
    if p.get("volume"):
        body += f"- **Об'єм:** {p['volume']}\n"
    if p.get("volume_form"):
        body += f"- **Форма випуску:** {p['volume_form']}\n"
    if p.get("mechanism_of_action"):
        body += f"\n## Механізм дії\n{p['mechanism_of_action']}\n"
    if p.get("duration_effect") or p.get("onset"):
        body += f"\n## Тривалість і початок дії\n"
        if p.get("duration_effect"):
            body += f"- **Тривалість ефекту:** {p['duration_effect']}\n"
        if p.get("onset"):
            body += f"- **Коли видно результат:** {p['onset']}\n"
    if p.get("indications"):
        body += f"\n## Показання\n{p['indications']}\n"
    if p.get("contraindications_absolute") or p.get("contraindications_relative") or p.get("pregnancy_lactation") or p.get("age_restrictions"):
        body += f"\n## Протипоказання\n"
        if p.get("contraindications_absolute"):
            body += f"**Абсолютні:** {p['contraindications_absolute']}\n\n"
        if p.get("contraindications_relative"):
            body += f"**Відносні:** {p['contraindications_relative']}\n\n"
        if p.get("pregnancy_lactation"):
            body += f"**Вагітність / лактація:** {p['pregnancy_lactation']}\n"
        if p.get("age_restrictions"):
            body += f"**Вікові обмеження:** {p['age_restrictions']}\n"
    has_protocol = any(p.get(k) for k in ["zones_allowed", "zones_forbidden", "injection_depth", "technique", "dosage", "course_count", "interval_repeat"])
    if has_protocol:
        body += f"\n## Протокол застосування\n"
        if p.get("zones_allowed"):
            body += f"**✅ Дозволені зони:** {p['zones_allowed']}\n\n"
        if p.get("zones_forbidden"):
            body += f"**⛔ ЗАБОРОНЕНІ зони:** {p['zones_forbidden']}\n\n"
        if p.get("injection_depth"):
            body += f"**Глибина введення:** {p['injection_depth']}\n"
        if p.get("technique"):
            body += f"**Техніка:** {p['technique']}\n"
        if p.get("dosage"):
            body += f"**Дозування за процедуру:** {p['dosage']}\n"
        if p.get("course_count"):
            body += f"**Кількість процедур у курсі:** {p['course_count']}\n"
        if p.get("interval_repeat"):
            body += f"**Інтервал між процедурами:** {p['interval_repeat']}\n"
    if p.get("compatibility_emet") or p.get("compatibility_devices"):
        body += f"\n## Сумісність\n"
        if p.get("compatibility_emet"):
            body += f"**З EMET-препаратами:** {p['compatibility_emet']}\n\n"
        if p.get("compatibility_devices"):
            body += f"**З апаратами (HIFU/RF/лазер):** {p['compatibility_devices']}\n"
    if p.get("recovery") or p.get("post_procedure_care") or p.get("side_effects_common"):
        body += f"\n## Реабілітація і догляд\n"
        if p.get("recovery"):
            body += f"**Down-time:** {p['recovery']}\n\n"
        if p.get("post_procedure_care"):
            body += f"**Догляд після процедури:** {p['post_procedure_care']}\n\n"
        if p.get("side_effects_common"):
            body += f"**Поширені побічні (норма):** {p['side_effects_common']}\n"
    if p.get("storage_temperature"):
        body += f"\n## Зберігання\n{p['storage_temperature']}\n"
    return yaml_header + body


def make_sales_md(p: dict) -> str:
    """Створює sales-картку з playbook."""
    canonical = canonical_for_rag(p["canonical"])
    safe_name = p["canonical"][:80]
    yaml_header = (
        f"---\n"
        f"product_name: {safe_name!r}\n"
        f"product_canonical: {canonical!r}\n"
        f"section: sales\n"
        f"source: sales_director_template_2026\n"
        f"---\n\n"
    )
    body = f"# {safe_name} — Sales Playbook\n\n"
    if p.get("price_segment"):
        body += f"**Ціновий сегмент:** {p['price_segment']}\n\n"
    if p.get("usp"):
        body += f"## USP (Unique Selling Proposition)\n{p['usp']}\n\n"
    if p.get("vs_competitors"):
        body += f"## VS конкуренти\n{p['vs_competitors']}\n\n"
    if p.get("common_objections"):
        body += f"## ТОП-3 заперечення лікарів + аргументи\n{p['common_objections']}\n\n"
    if p.get("killer_phrase"):
        body += f"## Killer phrase для лікаря\n«{p['killer_phrase']}»\n\n"
    if p.get("cross_sell"):
        body += f"## Cross-sell з EMET-препаратами\n{p['cross_sell']}\n\n"
    if p.get("ideal_patient"):
        body += f"## Профіль ідеального пацієнта\n{p['ideal_patient']}\n\n"
    if p.get("before_after_url"):
        body += f"## Кейси До/Після\n{p['before_after_url']}\n"
    return yaml_header + body


def slugify(name: str) -> str:
    """Безпечне ім'я файлу."""
    s = re.sub(r"[^\wЀ-ӿ\s-]", "", name).strip()
    s = re.sub(r"\s+", "_", s)
    return s[:60]


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, help="Path to filled Excel template")
    parser.add_argument("--output-dir", default="data/manual_product_cards",
                       help="Куди зберігати .md картки")
    parser.add_argument("--dry-run", action="store_true",
                       help="Лише надрукувати картки, не зберігати/не індексувати")
    parser.add_argument("--rebuild", action="store_true",
                       help="Перевідбудувати products_openai індекс (інакше тільки картки)")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"❌ Не знайдено {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    products = read_xlsx(xlsx_path)
    print(f"✅ Прочитано {len(products)} продуктів з {xlsx_path}")

    # Stats: скільки полів заповнено
    for p in products:
        filled = sum(1 for k, v in p.items() if v and str(v).strip())
        print(f"  {p['canonical'][:60]}: {filled}/{len(COLUMN_MAP)} полів")

    output_dir = Path(args.output_dir)
    if args.dry_run:
        # Print first product as preview
        if products:
            print("\n=== PREVIEW (clinical card for first product) ===\n")
            print(make_clinical_md(products[0]))
            print("\n=== PREVIEW (sales card for first product) ===\n")
            print(make_sales_md(products[0]))
        return

    output_dir.mkdir(parents=True, exist_ok=True)
    saved = 0
    for p in products:
        slug = slugify(p["canonical"])
        clinical_path = output_dir / f"{slug}__clinical.md"
        sales_path = output_dir / f"{slug}__sales.md"
        clinical_path.write_text(make_clinical_md(p), encoding="utf-8")
        sales_path.write_text(make_sales_md(p), encoding="utf-8")
        saved += 2
    print(f"\n✅ Збережено {saved} карток у {output_dir}/")

    if args.rebuild:
        # Тригернути перебудову products_openai через sync_manager
        # Реалізація в sync_manager — додати manual_product_cards у split
        print("\n🔄 Запускаю перебудову products_openai...")
        try:
            sys.path.insert(0, "/app")
            from sync_manager import _split_coach_to_products_competitors
            _split_coach_to_products_competitors()
            print("✅ Перебудова завершена")
        except Exception as e:
            print(f"❌ Помилка перебудови: {e}")
            print("Запусти вручну на сервері:")
            print('  docker exec emet_bot_app python -c "from sync_manager import _split_coach_to_products_competitors; _split_coach_to_products_competitors()"')


if __name__ == "__main__":
    main()
