"""tools/build_product_template_xlsx.py — будує Excel-шаблон для медвідділу.

Що робить:
1. Читає data/products_template_extracted.json (вже згенерований через RAG+LLM)
2. Створює Excel:
   - Лист 1 «Інструкція» — для медвідділу + Sales Director
   - Лист 2 «Продукти» — таблиця 14 продуктів × 29 полів, попередньо заповнена
   - Лист 3 «Глосарій» — пояснення складних полів
3. Зберігає docs/product_template_emet.xlsx

Запуск (локально):
    python tools/build_product_template_xlsx.py
"""
import json
import os
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


PROJECT_ROOT = Path(__file__).parent.parent
JSON_INPUT = PROJECT_ROOT / "data" / "products_template_v3.json"
XLSX_OUTPUT = PROJECT_ROOT / "docs" / "product_template_emet.xlsx"


# Лінк на «До/Після» — з main.py BEFORE_AFTER_LINKS (станом на 24.04)
BEFORE_AFTER_LINKS_DEFAULTS = {
    "Ellansé S": "https://links.emet.in.ua/ellanse-do-i-pislya/",
    "Ellansé M": "https://links.emet.in.ua/ellanse-do-i-pislya/",
    "Petaran": "https://links.emet.in.ua/protokol-zastosuvannya-petaran-poly-plla/",
    "Vitaran": "https://links.emet.in.ua/vitaran/case-vitaran/",
    "HP Cell Vitaran i": "https://links.emet.in.ua/vitaran/case-vitaran/",
    "HP Cell Vitaran iII": "https://links.emet.in.ua/vitaran/case-vitaran/",
    "HP Cell Vitaran Whitening": "https://links.emet.in.ua/vitaran/case-vitaran/",
    "HP Cell Vitaran Tox Eye": "https://links.emet.in.ua/vitaran/case-vitaran/",
    "EXOXE": "https://links.emet.in.ua/exoxe/case-exoxe/",
    "Neuramis": "https://links.emet.in.ua/neuramis_before_and_after/",
    "IUSE Hair Regrowth": "https://links.emet.in.ua/casey/",
    "IUSE Skinbooster HA20": "https://links.emet.in.ua/%d0%ba%d0%b5%d0%b9%d1%81%d0%b8-iuse-skin-boostercasey-2/",
}


# Колонки таблиці продуктів — порядок важливий
COLUMNS = [
    # IDENTIFICATION
    ("canonical", "Канонічна назва", 22, "med"),
    ("variants_writing", "Варіанти написання (UA+EN)", 28, "med"),
    ("manufacturer", "Виробник + країна", 28, "med"),
    ("category", "Категорія", 22, "med"),
    ("certification", "Сертифікація (CE / МОЗ / INCI)", 26, "med"),
    # COMPOSITION
    ("composition", "Активний склад + концентрація", 40, "med"),
    ("volume", "Об'єм упаковки", 18, "med"),
    ("volume_form", "Форма випуску", 22, "med"),
    ("mechanism_of_action", "Механізм дії", 40, "med"),
    # CLINICAL
    ("duration_effect", "Тривалість ефекту", 20, "med"),
    ("onset", "Коли видно результат", 22, "med"),
    ("indications", "Показання", 50, "med"),
    ("contraindications_absolute", "Абсолютні протипоказання", 50, "med"),
    ("contraindications_relative", "Відносні протипоказання", 50, "med"),
    ("pregnancy_lactation", "Вагітність / лактація", 24, "med"),
    ("age_restrictions", "Вікові обмеження", 18, "med"),
    # PROTOCOL
    ("zones_allowed", "Дозволені зони (✅)", 36, "med"),
    ("zones_forbidden", "ЗАБОРОНЕНІ зони (⛔)", 36, "med"),
    ("injection_depth", "Глибина введення", 24, "med"),
    ("technique", "Техніка", 30, "med"),
    ("dosage", "Дозування за процедуру", 26, "med"),
    ("course_count", "Кількість процедур у курсі", 20, "med"),
    ("interval_repeat", "Інтервал між процедурами", 26, "med"),
    # COMPATIBILITY
    ("compatibility_emet", "Сумісність з EMET-препаратами", 40, "med"),
    ("compatibility_devices", "Сумісність з апаратами (HIFU/RF/лазер)", 40, "med"),
    # AFTER-CARE
    ("recovery", "Реабілітація / down-time", 24, "med"),
    ("post_procedure_care", "Догляд ПІСЛЯ процедури", 36, "med"),
    ("side_effects_common", "Поширені побічні (норма)", 40, "med"),
    ("red_flags", "Червоні прапорці (терміново до медвідділу)", 40, "med"),
    # STORAGE
    ("storage_temperature", "Температура зберігання", 22, "med"),
    # SALES PLAYBOOK (Sales Director)
    ("price_segment", "Ціновий сегмент (бюджет/середній/преміум)", 22, "sales"),
    ("usp", "USP (1 речення — головна перевага)", 50, "sales"),
    ("vs_competitors", "VS конкуренти (sales-кут)", 50, "sales"),
    ("common_objections", "ТОП-3 заперечення лікарів + аргумент", 50, "sales"),
    ("killer_phrase", "Killer phrase (готова фраза для лікаря)", 40, "sales"),
    ("cross_sell", "З якими EMET-препаратами продається разом", 40, "sales"),
    ("ideal_patient", "Профіль ідеального пацієнта (psycho + клініка)", 40, "sales"),
    ("before_after_url", "Посилання на До/Після (links.emet.in.ua)", 40, "sales"),
]


HEADER_FILL_MED = PatternFill("solid", fgColor="066AAB")     # EMET синій — медвідділ
HEADER_FILL_SALES = PatternFill("solid", fgColor="C62828")   # червоний — sales playbook
HEADER_FONT = Font(name="Plus Jakarta Sans", size=10, bold=True, color="FFFFFF")
CELL_FONT = Font(name="Plus Jakarta Sans", size=10)
EMPTY_FILL = PatternFill("solid", fgColor="FFF3E0")           # оранжевий — треба заповнити
PREFILLED_FILL = PatternFill("solid", fgColor="E8F5E9")       # зелений — попередньо заповнено
CENTER_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CENTER_BOLD = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)


def add_instruction_sheet(wb: Workbook):
    ws = wb.create_sheet("📖 Інструкція", 0)
    ws.sheet_properties.tabColor = "066AAB"

    instructions = [
        ("EMET — Картки продуктів для бота", 18, True, "066AAB"),
        ("", 11, False, None),
        ("📋 Що це за документ", 14, True, "066AAB"),
        (
            "Цей файл — джерело правди про кожен продукт EMET для AI-бота.\n"
            "Бот використовує його щоб давати менеджерам точні відповіді про склад, "
            "протипоказання, протоколи, тривалість ефекту тощо.\n\n"
            "Зараз бот часто каже «не наведено в документах» — після заповнення цього файлу "
            "відповіді стануть точнішими.\n\n"
            "💡 У файлі НЕ всі продукти — ESSE-лінія обробляється окремо великою таблицею.",
            11, False, None
        ),
        ("", 11, False, None),
        ("👥 Хто заповнює (за кольором заголовків у таблиці)", 14, True, "066AAB"),
        (
            "🟦 СИНІ заголовки = МЕДВІДДІЛ — об'єктивні факти:\n"
            "    склад, показання, протипоказання, протокол, зберігання, сертифікація, механізм дії.\n\n"
            "🟥 ЧЕРВОНІ заголовки = SALES DIRECTOR — sales playbook:\n"
            "    USP, порівняння з конкурентами, заперечення, killer phrase, cross-sell,\n"
            "    ціновий сегмент, профіль пацієнта, посилання на До/Після.",
            11, False, None
        ),
        ("", 11, False, None),
        ("🎨 Кольори клітинок (що означають)", 14, True, "066AAB"),
        (
            "🟢 ЗЕЛЕНИЙ — попередньо заповнено автоматично з документів. ПЕРЕВІРИТИ і виправити якщо неточно.\n"
            "🟧 ОРАНЖЕВИЙ — поле порожнє, потрібно заповнити вручну.\n\n"
            "📝 Виправити зелене — просто переписуєш текст у клітинці. Не треба нічого позначати.\n"
            "📝 Якщо інформації нема ніде — лиши порожнім (краще пусте ніж вигадане).",
            11, False, None
        ),
        ("", 11, False, None),
        ("✅ Принципи заповнення (КРИТИЧНО)", 14, True, "066AAB"),
        (
            "1. КОРОТКО І КОНКРЕТНО — 1 факт = 1 пункт. Без води («революційна формула»).\n"
            "2. ЦИФРИ ЦИФРАМИ — «18 міс», не «вісімнадцять». «20 мг/мл», не «висока концентрація».\n"
            "3. ЯКЩО НЕ ЗНАЄШ — пропусти. Краще порожнє ніж вигадане.\n"
            "4. ПУНКТИ ЧЕРЕЗ ; (крапка з комою) — НЕ через , НЕ через нову лінію.\n"
            "   Приклад: «Вагітність; лактація; алергія на компоненти»\n"
            "5. ДЛЯ КОСМЕЦЕВТИКИ (Vitaran Skin Healer) — НЕ «CE», а «INCI».\n"
            "6. ВАРІАНТИ ПРОДУКТІВ — кожен у своєму рядку (Ellansé S окремо від M, кожен Neuramis окремо).\n"
            "7. ЦІНИ І МАРЖА — НЕ ВКАЗУВАТИ. Замість — поле «Ціновий сегмент»: бюджет / середній / преміум / люкс.",
            11, False, None
        ),
        ("", 11, False, None),
        ("⛔ Заборонені фрази (юр. контроль) — у всіх полях", 14, True, "C62828"),
        (
            "❌ «mass-market» / «мас-маркет»\n"
            "❌ «аптечний аналог»\n"
            "❌ «тестова процедура» / «привезти зразки» / «пробна партія»\n"
            "❌ «вічний філер» / «гарантований результат» / «безпечно для всіх»\n"
            "✅ Замість: «доступний / середній / преміум сегмент», «передбачуваний результат у клінічних дослідженнях»",
            11, False, None
        ),
        ("", 11, False, None),
        ("📞 Якщо незрозуміло", 14, True, "066AAB"),
        (
            "• Дивись лист «📚 Глосарій» (3-й вгорі) — там опис кожного поля з прикладами.\n"
            "• Якщо в зеленій клітинці помилка — переписуєш текст. Видалити — Delete.\n"
            "• Питання — пиши IT Director.",
            11, False, None
        ),
    ]

    row = 1
    for text, size, bold, color in instructions:
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(name="Plus Jakarta Sans", size=size, bold=bold,
                         color=color if color else "111827")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        # Автоматична висота рядка для довгих
        line_count = max(1, text.count("\n") + 1)
        ws.row_dimensions[row].height = max(20, 18 * line_count)
        row += 1

    ws.column_dimensions["A"].width = 100


def add_products_sheet(wb: Workbook, products_data):
    ws = wb.create_sheet("📦 Продукти", 1)
    ws.sheet_properties.tabColor = "2E7D32"

    # Header row — different colors for med (blue) vs sales (red) sections
    for col_idx, (key, label, width, owner) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = HEADER_FILL_SALES if owner == "sales" else HEADER_FILL_MED
        cell.font = HEADER_FONT
        cell.alignment = CENTER_BOLD
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Freeze header + first column (canonical name)
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 60

    # Data rows
    for row_idx, product in enumerate(products_data, 2):
        canonical = product.get("_canonical", "?")
        # Pre-fill before_after_url from defaults if empty
        if not product.get("before_after_url") and canonical in BEFORE_AFTER_LINKS_DEFAULTS:
            product["before_after_url"] = BEFORE_AFTER_LINKS_DEFAULTS[canonical]
        for col_idx, (key, label, width, owner) in enumerate(COLUMNS, 1):
            if key == "canonical":
                value = canonical
            else:
                value = product.get(key, "") or ""
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = CELL_FONT
            cell.alignment = CENTER_WRAP
            cell.border = THIN_BORDER
            # Color: green if pre-filled, orange if empty
            if value and value.strip():
                cell.fill = PREFILLED_FILL
            else:
                cell.fill = EMPTY_FILL

        # Висота рядка адаптивна — багато тексту = більше
        max_lines = 1
        for col_idx, (key, _, _, _) in enumerate(COLUMNS, 1):
            v = str(product.get(key, "") or "")
            lines_estimate = max(v.count("\n") + 1, len(v) // 40)
            max_lines = max(max_lines, lines_estimate)
        ws.row_dimensions[row_idx].height = min(300, max(40, 18 * max_lines))


def add_glossary_sheet(wb: Workbook):
    ws = wb.create_sheet("📚 Глосарій", 2)
    ws.sheet_properties.tabColor = "F57F17"

    glossary = [
        ("Поле", "Що писати", "Приклад"),
        ("Канонічна назва", "Офіційна назва — варіанти продукту окремими рядками", "Ellansé S (окремий рядок), Ellansé M (окремий рядок)"),
        ("Варіанти написання", "Усі можливі написання через кому (укр + англ + сленг менеджерів)",
         "Ellanse S, Елансе S, Эллансе S, Елансе-S"),
        ("Категорія",
         "Тип продукту: філер ГК / біостимулятор / ботулотоксин / екзосоми / нутрієнт / космецевтика тощо",
         "Колагено-стимулятор PCL"),
        ("Активний склад + концентрація",
         "Точна назва компонента + точна концентрація (мг/мл або %)",
         "PCL мікросфери 30% + КМЦ гель 70%"),
        ("Тривалість ефекту",
         "ТОЧНА цифра в місяцях/тижнях, без округлень",
         "18 місяців (для S), 24 місяці (для M)"),
        ("Показання",
         "3-5 пунктів через крапку з комою — для яких задач/зон/проблем",
         "Корекція об'ємних втрат середньої третини обличчя; контурна пластика підборіддя; відновлення скроневої зони"),
        ("Абсолютні протипоказання",
         "Стани коли НІКОЛИ не можна, через крапку з комою",
         "Вагітність та лактація; аутоімунні захворювання у фазі загострення; алергія на компоненти; активний інфекційний процес"),
        ("Відносні протипоказання",
         "Стани коли треба обережно (з обмеженнями)",
         "Прийом антикоагулянтів; перенесена герпетична інфекція"),
        ("Дозволені зони",
         "Куди МОЖНА вводити/застосовувати, з префіксом ✅",
         "✅ скроні; ✅ виличні зони; ✅ підборіддя; ✅ нижня щелепа"),
        ("ЗАБОРОНЕНІ зони",
         "Куди НЕ МОЖНА, з префіксом ⛔",
         "⛔ губи; ⛔ періорбітальна зона; ⛔ glabella"),
        ("Сумісність з EMET-препаратами",
         "З якими нашими препаратами можна комбінувати + інтервал. Назва препарату + дія",
         "Petaran: 6 міс в одну зону, різні зони — одна процедура; Neuramis: можна одразу різні шари"),
        ("Сумісність з апаратами",
         "HIFU / RF / лазер / мікронідлінг + інтервали. ✅ або ⛔",
         "✅ HIFU, RF — після 1-2 міс; ⛔ лазерна шліфовка — мінімум 3 міс"),
        ("Реабілітація / down-time",
         "Скільки днів пацієнт відновлюється",
         "3-7 днів (можливі набряки, гематоми)"),
        ("Поширені побічні (норма)",
         "Які реакції — це нормально, не привід панікувати",
         "Набряк 1-3 дні; гематома 3-7 днів; чутливість 2-4 тижні"),
        ("Червоні прапорці",
         "Симптоми коли ТЕРМІНОВО до медвідділу — для редиректу бота",
         "Тривала асиметрія >1 міс; нодульози; зміна кольору шкіри; ішемічні ознаки"),
        ("Сертифікація",
         "Для ін'єкційних: CE Mark + клас + МОЗ. Для космецевтики: INCI / COSMOS / ECOCERT (НЕ CE!)",
         "CE Mark клас III + МОЗ України № 12345"),
        ("USP", "Унікальна перевага в ОДНОМУ реченні (макс 30 слів). Конкретні факти, не вода.",
         "Ellansé — єдиний у портфелі EMET PCL-біостимулятор з клінічно передбачуваною тривалістю до 24 міс."),
        ("VS конкуренти",
         "Порівняння з 1-3 конкурентами по конкретних фактах (не «ми кращі»)",
         "VS Sculptra: миттєвий об'єм + стимуляція vs тільки стимуляція. 1 процедура vs 3-4."),
        ("ТОП-3 заперечення",
         "3 найчастіших заперечення лікарів + готовий аргумент-відповідь",
         "«Дорого»: 1 процедура замість 3-4 базових = вищий чек, менше слотів."),
        ("Профіль ідеального пацієнта",
         "Хто ідеальний клієнт — psycho + клініка (вік, стан, психотип, тригер покупки)",
         "40-55 років; втрата об'ємів; преміум-сегмент; цінує довговічність; не довіряє ін'єкціям спочатку"),
        ("Механізм дії",
         "Як працює на тканинному рівні (для відповідей бота на 'чому це працює')",
         "PCL-мікросфери стимулюють синтез колагену типу I; КМЦ дає миттєвий волюмізуючий ефект"),
        ("Догляд після процедури",
         "Що пацієнту НЕ можна / можна після процедури (окремо від down-time)",
         "24 год без макіяжу; 7 днів без сауни; 2 тижні без активного спорту"),
        ("Вагітність / лактація",
         "Так/ні + умови (топ-питання від лікарів)",
         "Протипоказано; абсолютно не застосовується; немає клінічних досліджень"),
        ("Вікові обмеження",
         "З якого віку дозволено / не рекомендовано після",
         "18+ років; не рекомендовано <30 років"),
        ("Ціновий сегмент",
         "Бюджет / середній / преміум / люкс — БЕЗ конкретних чисел!",
         "Преміум"),
        ("Killer phrase",
         "ОДНА коротка фраза яку менеджер скаже лікарю дослівно — закриває угоду",
         "«Один продукт замість трьох — економія слотів вашого графіку та вища маржа»"),
        ("Cross-sell",
         "З якими EMET-препаратами продається разом (комбо-протоколи)",
         "Petaran (різні зони, +30% до чеку); Vitaran i (об'єм + поверхневе омолодження); Neuramis (об'єм + філер у губи)"),
        ("Посилання на До/Після",
         "URL з links.emet.in.ua з кейсами — попередньо заповнено з бази бота",
         "https://links.emet.in.ua/ellanse-do-i-pislya/"),
    ]

    for row_idx, row_data in enumerate(glossary, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.fill = HEADER_FILL_MED
                cell.font = HEADER_FONT
                cell.alignment = CENTER_BOLD
            else:
                cell.font = CELL_FONT
                cell.alignment = CENTER_WRAP
            cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 60
    ws.row_dimensions[1].height = 30
    for r in range(2, len(glossary) + 1):
        ws.row_dimensions[r].height = 60


def main():
    if not JSON_INPUT.exists():
        print(f"❌ Не знайдено {JSON_INPUT}")
        return
    with open(JSON_INPUT, "r", encoding="utf-8") as f:
        products = json.load(f)

    wb = Workbook()
    # Видалити дефолтний пустий лист
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    add_instruction_sheet(wb)
    add_products_sheet(wb, products)
    add_glossary_sheet(wb)

    XLSX_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_OUTPUT)
    size_kb = XLSX_OUTPUT.stat().st_size / 1024
    print(f"✅ Збережено: {XLSX_OUTPUT} ({size_kb:.1f} KB)")
    print(f"   Продуктів: {len(products)}")
    print(f"   Колонок:   {len(COLUMNS)}")


if __name__ == "__main__":
    main()
