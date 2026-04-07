"""
make_course_template.py — генерирует Excel-шаблон курса для EMET-бота.
Запускать ЛОКАЛЬНО: python make_course_template.py
Создаёт файл course_template.xlsx рядом со скриптом.
"""
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

GREEN  = "1F7A4A"
LGRAY  = "F2F2F2"
YELLOW = "FFF2CC"
ORANGE = "FCE4D6"
WHITE  = "FFFFFF"
DGRAY  = "595959"

def hdr(ws, row, col, value, bg=GREEN, fg=WHITE, bold=True, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=fg, size=11)
    c.alignment = Alignment(wrap_text=wrap, vertical="center", horizontal="center")
    thin = Side(style="thin", color="AAAAAA")
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return c

def cell(ws, row, col, value="", bg=WHITE, bold=False, wrap=True, italic=False, fg="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=fg, size=10, italic=italic)
    c.alignment = Alignment(wrap_text=wrap, vertical="top")
    thin = Side(style="thin", color="CCCCCC")
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return c


wb = Workbook()

# ─────────────────────────────────────────────────
# ЛИСТ 1 — ІНСТРУКЦІЯ
# ─────────────────────────────────────────────────
ws0 = wb.active
ws0.title = "Інструкція"
ws0.column_dimensions["A"].width = 110

rows = [
    ("EMET Bot — Шаблон навчального курсу", GREEN, WHITE, True, 16),
    ("", WHITE, "000000", False, 11),
    ("ЯК ЗАПОВНИТИ", LGRAY, DGRAY, True, 12),
    ("", WHITE, "000000", False, 11),
    ("1.  Перейдіть на аркуш «Курс» — заповніть назву та опис курсу (рядок 2).", WHITE, "000000", False, 11),
    ("2.  Перейдіть на аркуш «Теми і тести» — кожен рядок = одне питання.", WHITE, "000000", False, 11),
    ("3.  Стовпці «Тема №» та «Назва теми» повторюйте для кожного питання одної теми.", WHITE, "000000", False, 11),
    ("4.  «Текст теми» (матеріал для навчання) заповнюйте ЛИШЕ в першому рядку кожної теми — далі залишайте порожнім.", WHITE, "000000", False, 11),
    ("5.  «Правильна відповідь» — введіть цифру: 1, 2, 3 або 4 (відповідає Варіанту 1..4).", WHITE, "000000", False, 11),
    ("6.  Мінімум 1 питання на тему, рекомендовано 3–5.", WHITE, "000000", False, 11),
    ("", WHITE, "000000", False, 11),
    ("ВАЖЛИВО", ORANGE, "000000", True, 12),
    ("", WHITE, "000000", False, 11),
    ("• Не змінюйте назви аркушів («Курс», «Теми і тести»).", WHITE, "000000", False, 11),
    ("• Не видаляйте рядок з заголовками (рядок 1 на аркуші «Теми і тести»).", WHITE, "000000", False, 11),
    ("• Текст теми підтримує Telegram Markdown: *жирний* _курсив_ - пункти.", WHITE, "000000", False, 11),
    ("• Кількість тем необмежена; питань у темі — необмежена.", WHITE, "000000", False, 11),
    ("", WHITE, "000000", False, 11),
    ("ЯК ЗАВАНТАЖИТИ В БОТ", LGRAY, DGRAY, True, 12),
    ("", WHITE, "000000", False, 11),
    ("Варіант A (рекомендовано): надішліть файл адміністратору — він запустить import_course.py на сервері.", WHITE, "000000", False, 11),
    ("Варіант B: адміністратор запускає в Telegram команду /upload_course і відправляє xlsx-файл прямо в чат боту.", WHITE, "000000", False, 11),
    ("", WHITE, "000000", False, 11),
    ("ПРИКЛАД ЗАПОВНЕННЯ — дивіться аркуш «Приклад»", YELLOW, "000000", True, 11),
]

for i, (text, bg, fg, bold, size) in enumerate(rows, 1):
    c = ws0.cell(row=i, column=1, value=text)
    c.fill = PatternFill("solid", fgColor=bg)
    c.font = Font(bold=bold, color=fg, size=size)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws0.row_dimensions[i].height = 22 if text else 8

ws0.sheet_view.showGridLines = False


# ─────────────────────────────────────────────────
# ЛИСТ 2 — КУРС (мета-дані)
# ─────────────────────────────────────────────────
ws1 = wb.create_sheet("Курс")
ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 80

hdr(ws1, 1, 1, "Поле")
hdr(ws1, 1, 2, "Значення")

for r, (label, placeholder) in enumerate([
    ("Назва курсу",  "Ellansé — базовий курс продажів"),
    ("Опис курсу",   "Склад, показання, техніки продажу та заперечення для препарату Ellansé"),
], 2):
    cell(ws1, r, 1, label, bg=LGRAY, bold=True, wrap=False)
    cell(ws1, r, 2, placeholder, bg=YELLOW, italic=True)
    ws1.row_dimensions[r].height = 24

ws1.sheet_view.showGridLines = False


# ─────────────────────────────────────────────────
# ЛИСТ 3 — ТЕМИ І ТЕСТИ
# ─────────────────────────────────────────────────
ws2 = wb.create_sheet("Теми і тести")

COL_WIDTHS = [9, 30, 60, 60, 28, 28, 28, 28, 18]
COL_LABELS = [
    "Тема №",
    "Назва теми",
    "Текст теми\n(матеріал для навчання)",
    "Питання тесту",
    "Варіант 1",
    "Варіант 2",
    "Варіант 3",
    "Варіант 4",
    "Правильна\nвідповідь (1–4)",
]

for ci, (w, label) in enumerate(zip(COL_WIDTHS, COL_LABELS), 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w
    hdr(ws2, 1, ci, label, wrap=True)
ws2.row_dimensions[1].height = 34

# Валидация: правильный ответ только 1-4
dv = DataValidation(type="whole", operator="between", formula1="1", formula2="4",
                    error="Введіть 1, 2, 3 або 4", errorTitle="Помилка",
                    prompt="Номер правильного варіанту (1–4)", promptTitle="Правильна відповідь")
dv.sqref = "I2:I1000"
ws2.add_data_validation(dv)

ws2.freeze_panes = "A2"
ws2.sheet_view.showGridLines = True

# Пустые строки для заполнения (30 строк)
for r in range(2, 32):
    bg = LGRAY if r % 2 == 0 else WHITE
    for ci in range(1, 10):
        cell(ws2, r, ci, bg=bg)
    ws2.row_dimensions[r].height = 48


# ─────────────────────────────────────────────────
# ЛИСТ 4 — ПРИКЛАД (заполненный образец)
# ─────────────────────────────────────────────────
ws3 = wb.create_sheet("Приклад")

for ci, (w, label) in enumerate(zip(COL_WIDTHS, COL_LABELS), 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w
    hdr(ws3, 1, ci, label, wrap=True)
ws3.row_dimensions[1].height = 34

example_rows = [
    # (тема_№, назва_теми, текст_теми, питання, opt1, opt2, opt3, opt4, правильна)
    (
        1,
        "Що таке Ellansé?",
        "*Ellansé* — ін'єкційний філер на основі *PCL* (полікапролактон).\n\n"
        "PCL — біодеградуючий полімер, який стимулює вироблення власного колагену.\n\n"
        "*Відмінність від ГК-філерів:*\n"
        "- ГК дає об'єм на 6–12 міс, потім розсмоктується повністю\n"
        "- Ellansé: ефект 2–4 роки, замінюється власним колагеном пацієнта\n\n"
        "*Лінійка:* S (1 рік), M (2 роки), L (3 роки), E (4 роки)",
        "Яка основна діюча речовина Ellansé?",
        "Гіалуронова кислота",
        "Полікапролактон (PCL)",
        "PDRN",
        "Ботулотоксин",
        2,
    ),
    (
        1, "Що таке Ellansé?", "",
        "Скільки років тримається ефект лінійки Ellansé E?",
        "1 рік", "2 роки", "3 роки", "4 роки",
        4,
    ),
    (
        1, "Що таке Ellansé?", "",
        "Чим Ellansé відрізняється від ГК-філерів?",
        "Він дешевший",
        "Він стимулює вироблення власного колагену і тримається 2–4 роки",
        "Він не потребує введення лікарем",
        "Він розсмоктується за 3 місяці",
        2,
    ),
    (
        2,
        "Продажі через цінність",
        "*Формула:* ВЛАСТИВІСТЬ → ПЕРЕВАГА → ВИГОДА\n\n"
        "Приклад: 'PCL стимулює колаген (властивість) → ефект 2–4 роки (перевага) → пацієнт повертається рідше, але платить більше за процедуру (вигода для лікаря)'\n\n"
        "*Заперечення 'Дорого':*\n"
        "→ 'Лікар, ГК вимагає повтор кожні 9–12 місяців. Ellansé — раз на 2–4 роки. "
        "Пацієнт витрачає менше в сумі, а клініка отримує більший разовий чек і лояльного клієнта.'\n\n"
        "*Заперечення 'Є постачальник Juvederm':*\n"
        "→ 'Juvederm — чудовий ГК-філер для об'єму. Ellansé вирішує інше завдання — "
        "омолодження через власний колаген. Вони не замінюють, а доповнюють одне одного.'",
        "Яку основну вигоду для лікаря підкреслює Ellansé?",
        "Низька ціна препарату",
        "Мінімальна кількість повторних процедур = вищий разовий чек",
        "Простота введення",
        "Підходить для всіх пацієнтів без винятків",
        2,
    ),
    (
        2, "Продажі через цінність", "",
        "Як відповісти на заперечення 'Є постачальник Juvederm'?",
        "Сказати що Juvederm гірший",
        "Запропонувати знижку",
        "Пояснити що Ellansé і ГК-філери вирішують різні задачі і доповнюють одне одного",
        "Поступитись і не продавати",
        3,
    ),
]

for r_i, row_data in enumerate(example_rows, 2):
    bg = "#E8F5E9" if row_data[0] % 2 == 1 else "#E3F2FD"
    for ci, val in enumerate(row_data, 1):
        cell(ws3, r_i, ci, val, bg=bg.replace("#",""), wrap=True)
    ws3.row_dimensions[r_i].height = 80

ws3.freeze_panes = "A2"
ws3.sheet_view.showGridLines = True


# ─────────────────────────────────────────────────
# Сохранить
# ─────────────────────────────────────────────────
out = "course_template.xlsx"
wb.save(out)
print(f"Template saved: {out}")
