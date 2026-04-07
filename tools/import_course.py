"""
import_course.py — импортирует курс из Excel-шаблона в PostgreSQL бота.

Запуск (в Docker-контейнере):
  docker cp course_Ellanse.xlsx emet_bot_app:/app/course.xlsx
  docker exec -it emet_bot_app python import_course.py /app/course.xlsx

Или локально (если PostgreSQL доступен напрямую):
  python import_course.py course_Ellanse.xlsx
"""

import sys
import os
from dotenv import load_dotenv
load_dotenv()

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

import db  # EMET db.py — PostgreSQL connection pool


def import_course(path: str):
    wb = load_workbook(path, data_only=True)

    # --- Валидация структуры ---
    required_sheets = {"Курс", "Теми і тести"}
    missing = required_sheets - set(wb.sheetnames)
    if missing:
        print(f"ERROR: Missing sheets: {missing}")
        print(f"  Available: {wb.sheetnames}")
        sys.exit(1)

    # --- Читаем мета-данные курса ---
    ws_meta = wb["Курс"]
    title = str(ws_meta.cell(2, 2).value or "").strip()
    description = str(ws_meta.cell(3, 2).value or "").strip()

    if not title:
        print("ERROR: Course title is empty (sheet 'Курс', cell B2)")
        sys.exit(1)

    # --- Читаем темы и вопросы ---
    ws = wb["Теми і тести"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    topics = {}       # topic_num → {title, content, questions: [...]}
    topic_order = []  # сохраняем порядок появления тем

    for row_i, row in enumerate(rows, 2):
        if all(v is None or str(v).strip() == "" for v in row):
            continue  # пустая строка — пропускаем

        try:
            topic_num   = int(row[0]) if row[0] is not None else None
            topic_title = str(row[1] or "").strip()
            topic_text  = str(row[2] or "").strip()
            q_text      = str(row[3] or "").strip()
            opt1        = str(row[4] or "").strip()
            opt2        = str(row[5] or "").strip()
            opt3        = str(row[6] or "").strip()
            opt4        = str(row[7] or "").strip()
            correct_raw = row[8]
        except Exception as e:
            print(f"  WARNING: row {row_i} parse error: {e} — skipping")
            continue

        if topic_num is None:
            print(f"  WARNING: row {row_i} — 'Тема №' is empty — skipping")
            continue

        # Инициализируем тему при первом появлении
        if topic_num not in topics:
            if not topic_title:
                print(f"  WARNING: row {row_i} — topic {topic_num} has no title — skipping")
                continue
            topics[topic_num] = {
                "title": topic_title,
                "content": topic_text,
                "questions": [],
            }
            topic_order.append(topic_num)
        else:
            # Дополнительная строка той же темы — обновляем текст если он появился
            if topic_text and not topics[topic_num]["content"]:
                topics[topic_num]["content"] = topic_text

        # Добавляем вопрос (если заполнен)
        if q_text:
            options = [o for o in [opt1, opt2, opt3, opt4] if o]
            if len(options) < 2:
                print(f"  WARNING: row {row_i} — question '{q_text[:40]}...' has < 2 options — skipping question")
                continue

            try:
                correct_idx = int(correct_raw) - 1  # 0-indexed
            except (TypeError, ValueError):
                print(f"  WARNING: row {row_i} — correct answer '{correct_raw}' is not a number — skipping question")
                continue

            if correct_idx < 0 or correct_idx >= len(options):
                print(f"  WARNING: row {row_i} — correct answer {correct_idx+1} out of range (1–{len(options)}) — skipping")
                continue

            topics[topic_num]["questions"].append({
                "text": q_text,
                "options": [(opt, i == correct_idx) for i, opt in enumerate(options)],
            })

    if not topics:
        print("ERROR: No topics found. Check sheet 'Теми і тести'")
        sys.exit(1)

    # Предпросмотр
    print(f"\nCourse:      {title}")
    print(f"Description: {description or '(empty)'}")
    print(f"Topics:      {len(topics)}")
    total_q = sum(len(t["questions"]) for t in topics.values())
    print(f"Questions:   {total_q}")
    for tn in topic_order:
        t = topics[tn]
        print(f"  Topic {tn}: {t['title']} — {len(t['questions'])} questions")

    confirm = input("\nImport to database? [y/N] ").strip().lower()
    if confirm != "y":
        print("Cancelled.")
        sys.exit(0)

    # --- Запись в БД ---
    from datetime import datetime

    with db.get_connection() as conn:
        with conn.cursor() as cur:
            # Проверяем дубликат
            cur.execute("SELECT id FROM courses WHERE title=%s", (title,))
            existing = cur.fetchone()
            if existing:
                overwrite = input(f"\nCourse '{title}' already exists (id={existing[0]}). Overwrite? [y/N] ").strip().lower()
                if overwrite != "y":
                    print("Cancelled.")
                    sys.exit(0)
                # Удаляем каскадно
                cur.execute(
                    "DELETE FROM answer_options WHERE question_id IN "
                    "(SELECT q.id FROM questions q JOIN topics t ON q.topic_id=t.id WHERE t.course_id=%s)",
                    (existing[0],)
                )
                cur.execute(
                    "DELETE FROM questions WHERE topic_id IN (SELECT id FROM topics WHERE course_id=%s)",
                    (existing[0],)
                )
                cur.execute("DELETE FROM topics WHERE course_id=%s", (existing[0],))
                cur.execute("DELETE FROM courses WHERE id=%s", (existing[0],))
                print(f"  Old course deleted.")

            # Вставляем курс
            cur.execute(
                "INSERT INTO courses (title, description, created_at) VALUES (%s, %s, %s) RETURNING id",
                (title, description, datetime.now().isoformat())
            )
            course_id = cur.fetchone()[0]

            # Вставляем темы и вопросы
            for order, tn in enumerate(topic_order, 1):
                t = topics[tn]
                cur.execute(
                    "INSERT INTO topics (course_id, order_num, title, content) VALUES (%s,%s,%s,%s) RETURNING id",
                    (course_id, order, t["title"], t["content"])
                )
                topic_id = cur.fetchone()[0]

                for q in t["questions"]:
                    cur.execute(
                        "INSERT INTO questions (topic_id, text) VALUES (%s,%s) RETURNING id",
                        (topic_id, q["text"])
                    )
                    q_id = cur.fetchone()[0]
                    for opt_text, is_correct in q["options"]:
                        cur.execute(
                            "INSERT INTO answer_options (question_id, text, is_correct) VALUES (%s,%s,%s)",
                            (q_id, opt_text, 1 if is_correct else 0)
                        )

    print(f"\nDone! Course '{title}' imported as id={course_id}")
    print(f"Topics: {len(topic_order)}, Questions: {total_q}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python import_course.py <path_to_xlsx>")
        print("Example: python import_course.py course_Ellanse.xlsx")
        sys.exit(1)

    import_course(sys.argv[1])
