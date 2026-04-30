"""tools/import_esse_assortment_to_rag.py — імпорт повної ESSE-таблиці в RAG.

Структура ESSE_асортимент (роздріб).xlsx:
- 6 листів: ESSE CORE, ESSE SENSITIVE, ESSE PLUS, Консилери, Сонцезахисні, Набори
- Стандартні колонки на лист (різна кількість, але базові ті ж):
  Найменування | Показання | Склад | Країна виробник | Спосіб застосування |
  Об'єм | Опис | Про продукцію ESSE | Посилання на зображення | Посилання на товар

Цей скрипт:
1. Читає всі 6 листів
2. По кожному продукту генерує одну .md картку з YAML frontmatter
3. Зберігає у data/manual_product_cards/ (там же, де картки з основного шаблону)
4. Картки переживають auto-sync завдяки sync_manager patch'у

Запуск:
    python tools/import_esse_assortment_to_rag.py --xlsx "docs/ESSE_асортимент (роздріб).xlsx"
"""
import argparse
import re
import sys
from pathlib import Path
from openpyxl import load_workbook


# Назви канонічних ліній ESSE для metadata.product_canonical
LINE_TO_CANONICAL = {
    "Лінійка ЕSSE CORE": "ESSE",
    "Лінійка ESSE SENSITIVE": "ESSE",
    "Лінійка ESSE PLUS": "ESSE",
    "Консилери": "ESSE",
    "Сонцезахисні креми і тональні о": "ESSE",
    "Набори": "ESSE",
}

LINE_TO_SUBLINE = {
    "Лінійка ЕSSE CORE": "Esse Core",
    "Лінійка ESSE SENSITIVE": "Esse Sensitive",
    "Лінійка ESSE PLUS": "Esse Plus",
    "Консилери": "Esse Concealers",
    "Сонцезахисні креми і тональні о": "Esse Sun Protection & Foundation",
    "Набори": "Esse Sets",
}


def slugify(name: str) -> str:
    s = re.sub(r"[^\wЀ-ӿ\s-]", "", name).strip()
    s = re.sub(r"\s+", "_", s)
    return s[:80]


def make_esse_card(product: dict, line_name: str) -> str:
    """Картка одного ESSE-продукту."""
    name = product.get("name", "").strip()
    if not name:
        return ""
    safe_name = name[:120]
    subline = LINE_TO_SUBLINE.get(line_name, "Esse")
    yaml_header = (
        f"---\n"
        f"product_name: {safe_name!r}\n"
        f"product_canonical: 'ESSE'\n"
        f"product_subline: {subline!r}\n"
        f"section: clinical\n"
        f"source: esse_assortment_2026\n"
        f"---\n\n"
    )
    body = f"# {safe_name} — {subline}\n\n"
    if product.get("indications"):
        body += f"## Показання\n{product['indications']}\n\n"
    if product.get("composition"):
        body += f"## Склад\n{product['composition']}\n\n"
    if product.get("manufacturer"):
        body += f"**Виробник:** {product['manufacturer']}\n"
    if product.get("volume"):
        body += f"**Об'єм:** {product['volume']}\n\n"
    if product.get("how_to_use"):
        body += f"## Спосіб застосування\n{product['how_to_use']}\n\n"
    if product.get("description"):
        body += f"## Опис\n{product['description']}\n\n"
    if product.get("about_esse"):
        body += f"## Про продукцію ESSE\n{product['about_esse']}\n\n"
    if product.get("image_url"):
        body += f"**Зображення:** {product['image_url']}\n"
    if product.get("product_url"):
        body += f"**Купити:** {product['product_url']}\n"
    return yaml_header + body


# Маппінг колонок (назви заголовків можуть варіюватися між листами)
COLUMN_KEYS = {
    "name": ["найменування"],
    "indications": ["показання"],
    "composition": ["склад"],
    "manufacturer": ["країна виробник"],
    "how_to_use": ["спосіб застосування", "спосіб застосування:"],
    "volume": ["об'єм"],
    "description": ["опис"],
    "about_esse": ["про продукцію esse"],
    "image_url": ["посилання на зображення"],
    "product_url": ["посилання на товар"],
}


def find_col_idx(headers: list, header_keys: list) -> int | None:
    """Знайти індекс колонки за можливими ім'ями (case-insensitive)."""
    norm = [(h or "").strip().lower() for h in headers]
    for key in header_keys:
        if key in norm:
            return norm.index(key) + 1  # openpyxl 1-indexed
    return None


def read_esse_xlsx(xlsx_path: Path) -> list[dict]:
    """Читає всі 6 листів і повертає список продуктів з полями."""
    wb = load_workbook(xlsx_path, data_only=True)
    products = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Headers — рядок 1, очищені
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        col_idx = {k: find_col_idx(headers, v) for k, v in COLUMN_KEYS.items()}
        # Data rows — починаються з 2
        for r in range(2, ws.max_row + 1):
            name_val = ws.cell(row=r, column=col_idx["name"] or 1).value
            if not name_val or not str(name_val).strip():
                continue
            product = {"line": sheet_name}
            for key, idx in col_idx.items():
                if idx is None:
                    continue
                val = ws.cell(row=r, column=idx).value
                product[key] = (str(val).strip() if val is not None else "")
            products.append(product)
    return products


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx", required=True, help="Path to ESSE_асортимент xlsx")
    parser.add_argument("--output-dir", default="data/manual_product_cards")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        print(f"❌ Не знайдено {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    products = read_esse_xlsx(xlsx_path)
    print(f"✅ Прочитано {len(products)} ESSE-продуктів")

    # Stats per line
    by_line = {}
    for p in products:
        by_line.setdefault(p["line"], 0)
        by_line[p["line"]] += 1
    for line, cnt in by_line.items():
        print(f"  {line}: {cnt}")

    if args.dry_run:
        if products:
            print("\n=== PREVIEW (first product) ===\n")
            print(make_esse_card(products[0], products[0]["line"]))
        return

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    saved = 0
    for p in products:
        # Скіп якщо нема назви або занадто мало даних
        filled = sum(1 for k in ["indications", "composition", "how_to_use", "description"] if p.get(k))
        if filled < 2:
            continue
        slug = slugify(p["name"])
        # Префікс ESSE_ + слаг ім'я
        path = output_dir / f"ESSE_{slug}.md"
        path.write_text(make_esse_card(p, p["line"]), encoding="utf-8")
        saved += 1
    print(f"\n✅ Збережено {saved} ESSE-карток у {output_dir}/")


if __name__ == "__main__":
    main()
