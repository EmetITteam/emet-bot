"""
One-off import: Тест Еllanse.xlsx → PostgreSQL
Запускать ЛОКАЛЬНО: python _import_ellanse.py
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

os.chdir(os.path.dirname(os.path.abspath(__file__)))
from dotenv import load_dotenv
load_dotenv()

from openpyxl import load_workbook
from datetime import datetime
import re
import db

XLSX = "Тест Еllanse.xlsx"

wb = load_workbook(XLSX, data_only=True)
print("Sheets:", wb.sheetnames)

# Лист с данными — "Ellanse" или "Теми і тести"
data_sheet = None
for candidate in ["Ellanse", "Теми і тести"]:
    if candidate in wb.sheetnames:
        data_sheet = candidate
        break

if data_sheet is None:
    print("ERROR: no data sheet found")
    sys.exit(1)
print("Reading data from sheet:", data_sheet)

# Мета-данные курса
ws_meta = wb["Курс"]
title = str(ws_meta.cell(2, 2).value or "").strip()
description = str(ws_meta.cell(3, 2).value or "").strip()
print(f"Course title: {title}")
print(f"Description:  {description}")

# Парсим темы и вопросы
ws = wb[data_sheet]
topics = {}       # topic_num(int) → {title, content, questions}
topic_order = []
last_topic_num = None

for row_i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    if all(v is None or str(v).strip() == "" for v in row):
        continue

    # Тема №  — может быть "4." или "" (наследует предыдущую)
    raw_num = str(row[0] or "").strip().rstrip(".")
    if raw_num.isdigit():
        topic_num = int(raw_num)
        last_topic_num = topic_num
    elif last_topic_num is not None:
        topic_num = last_topic_num  # пустой → продолжение предыдущей темы
    else:
        print(f"  WARNING row {row_i}: topic_num empty, no previous topic — skipping")
        continue

    topic_title = str(row[1] or "").strip().rstrip(",")
    topic_text  = str(row[2] or "").strip()
    q_text      = str(row[3] or "").strip()
    opt1        = str(row[4] or "").strip()
    opt2        = str(row[5] or "").strip()
    opt3        = str(row[6] or "").strip()
    opt4        = str(row[7] or "").strip()
    correct_raw = str(row[8] or "").strip()

    # Инициализация темы
    if topic_num not in topics:
        if not topic_title:
            print(f"  WARNING row {row_i}: topic {topic_num} no title — skipping")
            continue
        topics[topic_num] = {"title": topic_title, "content": "", "questions": []}
        topic_order.append(topic_num)

    # Текст темы — склеиваем все куски
    if topic_text:
        if topics[topic_num]["content"]:
            topics[topic_num]["content"] += "\n\n" + topic_text
        else:
            topics[topic_num]["content"] = topic_text

    # Вопрос
    if q_text:
        options = [o for o in [opt1, opt2, opt3, opt4] if o]
        if len(options) < 2:
            print(f"  WARNING row {row_i}: < 2 options for question — skipping")
            continue

        raw_correct = re.sub(r"[^0-9]", "", correct_raw)
        if not raw_correct:
            print(f"  WARNING row {row_i}: correct answer '{correct_raw}' invalid — skipping question")
            continue
        correct_idx = int(raw_correct) - 1  # 0-indexed

        if correct_idx < 0 or correct_idx >= len(options):
            print(f"  WARNING row {row_i}: correct_idx {correct_idx} out of range — skipping")
            continue

        topics[topic_num]["questions"].append({
            "text": q_text,
            "options": [(opt, i == correct_idx) for i, opt in enumerate(options)],
        })

# Превью
print(f"\n{'='*55}")
print(f"Course:    {title}")
print(f"Topics:    {len(topics)}")
total_q = sum(len(t['questions']) for t in topics.values())
print(f"Questions: {total_q}")
for tn in topic_order:
    t = topics[tn]
    print(f"  Topic {tn}: {t['title'][:50]} — {len(t['questions'])} q")
print(f"{'='*55}\n")

if not topics:
    print("ERROR: no topics parsed")
    sys.exit(1)

confirm = input("Import to DB? [y/N] ").strip().lower()
if confirm != "y":
    print("Cancelled.")
    sys.exit(0)

# Запись в БД
with db.get_connection() as conn:
    with conn.cursor() as cur:
        # Проверяем дубликат
        cur.execute("SELECT id FROM courses WHERE title=%s", (title,))
        existing = cur.fetchone()
        if existing:
            ow = input(f"Course '{title}' already exists (id={existing[0]}). Overwrite? [y/N] ").strip().lower()
            if ow != "y":
                print("Cancelled.")
                sys.exit(0)
            cur.execute(
                "DELETE FROM answer_options WHERE question_id IN "
                "(SELECT q.id FROM questions q JOIN topics t ON q.topic_id=t.id WHERE t.course_id=%s)",
                (existing[0],)
            )
            cur.execute("DELETE FROM questions WHERE topic_id IN (SELECT id FROM topics WHERE course_id=%s)", (existing[0],))
            cur.execute("DELETE FROM topics WHERE course_id=%s", (existing[0],))
            cur.execute("DELETE FROM courses WHERE id=%s", (existing[0],))
            print("  Old course removed.")

        cur.execute(
            "INSERT INTO courses (title, description, created_at) VALUES (%s,%s,%s) RETURNING id",
            (title, description, datetime.now().isoformat())
        )
        course_id = cur.fetchone()[0]

        for order, tn in enumerate(topic_order, 1):
            t = topics[tn]
            cur.execute(
                "INSERT INTO topics (course_id, order_num, title, content) VALUES (%s,%s,%s,%s) RETURNING id",
                (course_id, order, t["title"], t["content"])
            )
            topic_id = cur.fetchone()[0]

            for q in t["questions"]:
                cur.execute("INSERT INTO questions (topic_id, text) VALUES (%s,%s) RETURNING id", (topic_id, q["text"]))
                q_id = cur.fetchone()[0]
                for opt_text, is_correct in q["options"]:
                    cur.execute(
                        "INSERT INTO answer_options (question_id, text, is_correct) VALUES (%s,%s,%s)",
                        (q_id, opt_text, 1 if is_correct else 0)
                    )

print(f"\nDone! course_id={course_id}, topics={len(topic_order)}, questions={total_q}")
