import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

"""
Генерує Excel-шаблон synonyms_template.xlsx — локальний словник синонімів.
Запуск: python tests/create_synonyms_template.py
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = os.path.join(os.path.dirname(__file__), "synonyms_template.xlsx")

COLOR_HEADER  = "1F4E79"
COLOR_KB      = "DEEAF1"
COLOR_COACH   = "E2EFDA"
COLOR_CERTS   = "FFF2CC"
COLOR_COMBO   = "FCE4D6"

EXAMPLES = [
    # (тема, категорія, оригінальний_термін, синоніми_через_кому, мова, пояснення)
    # --- KB: HR ---
    ("HR / Звільнення", "kb",
     "увольнение",
     "звільнення, обходной лист, розрахунок, припинення трудового договору, звільнити, розрахувати",
     "RU+UA",
     "Менеджери пишуть по-різному — потрібно покрити обидві мови"),

    ("HR / Відпустка", "kb",
     "отпуск",
     "відпустка, відпустки, відпустку, щорічна відпустка, заява на відпустку, відгул",
     "RU+UA",
     "Різні відмінки і форми слова"),

    ("HR / Лікарняний", "kb",
     "больничный",
     "лікарняний, лікарняний лист, лист непрацездатності, бюлетень, захворів",
     "RU+UA",
     ""),

    ("IT / CRM", "kb",
     "CRM",
     "срм, crm-система, клієнтська база, база клієнтів, доступ до crm",
     "EN+UA",
     "CRM пишуть і великими і малими буквами"),

    ("Структура / Відділи", "kb",
     "отдел продаж",
     "комерційний відділ, sales відділ, відділ збуту, менеджери з продажу",
     "RU+UA",
     ""),

    # --- COACH: Препарати ---
    ("Препарат / Vitaran", "coach",
     "Vitaran",
     "Вітаран, витаран, PDRN, полідезоксирибонуклеотид, поліднк, Vitaran 2%",
     "UA+RU+EN",
     "Назва пишеться по-різному: Vitaran / Вітаран / витаран"),

    ("Препарат / Ellansé", "coach",
     "Ellansé",
     "Елансе, еланс, ellanse, PCL, полікапролактон, Ellanse S, Ellanse M",
     "UA+RU+EN",
     "Різні серії препарату"),

    ("Препарат / Neuramis", "coach",
     "Neuramis",
     "Нейраміс, нейрамис, Neuramis deep, Neuramis volume, філлер Neuramis",
     "UA+RU+EN",
     ""),

    ("Заперечення / Ціна", "coach",
     "дорого",
     "дорогой, дорого, висока ціна, коштує багато, не можу дозволити, бюджет, знижка",
     "RU+UA",
     "Всі варіанти як лікар може сказати що дорого"),

    ("Заперечення / Конкурент", "coach",
     "аналог",
     "конкурент, аналог, замінник, дешевший, є інший, є краще, є кращий препарат",
     "RU+UA",
     ""),

    # --- CERTS: Сертифікати ---
    ("Сертифікати", "certs",
     "сертификат",
     "сертифікат, реєстраційне посвідчення, дозвіл, ліцензія, документи на препарат, реєстрація",
     "RU+UA",
     "Різні назви одного й того ж документа"),

    ("Сертифікати / Vitaran", "certs",
     "сертификат Vitaran",
     "сертифікат Вітаран, документи на Витаран, реєстрація Vitaran, посвідчення Vitaran",
     "RU+UA",
     "Конкретний препарат + тип документа"),

    # --- COMBO ---
    ("Комбо / Поєднання", "combo",
     "комбинировать",
     "комбо, комбінувати, комбінація, поєднати, поєднання, сочетать, сочетание, combo протокол",
     "RU+UA+EN",
     "Всі варіанти як питають про комбо-протоколи"),
]

COLUMNS = [
    ("Тема / Категорія запиту", 28),
    ("Розділ бази",              12),
    ("Ключовий термін",          22),
    ("Синоніми (через кому)",    55),
    ("Мови",                     12),
    ("Коментар",                 35),
]

CAT_COLORS = {"kb": COLOR_KB, "coach": COLOR_COACH, "certs": COLOR_CERTS, "combo": COLOR_COMBO}
CAT_LABELS = {"kb": "🔍 KB", "coach": "💼 Коуч", "certs": "📜 Серт", "combo": "🔗 Комбо"}


def thin():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)


def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Synonyms"

    # Заголовок
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = "EMET Bot — Локальний словник синонімів для RAG-пошуку"
    c.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Інструкція
    ws.merge_cells("A2:F2")
    inst = ws["A2"]
    inst.value = (
        "Цей файл описує групи синонімів для локальної заміни prepare_search_query (без API-виклику). "
        "Колонка 'Синоніми' — все що бот додасть до пошукового запиту при знаходженні ключового терміну."
    )
    inst.font = Font(name="Calibri", italic=True, size=10, color="555555")
    inst.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 28

    # Заголовки колонок
    hfill = PatternFill("solid", fgColor="2E75B6")
    for ci, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=3, column=ci, value=name)
        cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin()
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[3].height = 24

    # Дані
    for ri, (topic, cat, term, synonyms, langs, comment) in enumerate(EXAMPLES, 4):
        fill = PatternFill("solid", fgColor=CAT_COLORS[cat])
        for ci, val in enumerate([topic, CAT_LABELS[cat], term, synonyms, langs, comment], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin()
        ws.row_dimensions[ri].height = 45

    # Порожні рядки для заповнення
    for i in range(5):
        ri = len(EXAMPLES) + 4 + i
        cat = ["kb", "coach", "certs", "combo", "kb"][i]
        fill = PatternFill("solid", fgColor=CAT_COLORS[cat])
        for ci, val in enumerate(["← тема", CAT_LABELS[cat], "← термін", "← синонім1, синонім2, ...", "UA+RU", "← коментар"], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.font = Font(name="Calibri", size=10, color="999999", italic=True)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin()
        ws.row_dimensions[ri].height = 45

    ws.freeze_panes = "A4"
    wb.save(OUTPUT)
    print(f"✅ Шаблон збережено: {OUTPUT}")


if __name__ == "__main__":
    build()
