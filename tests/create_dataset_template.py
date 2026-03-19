import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

"""
Генерує Excel-шаблон golden_dataset_template.xlsx для заповнення командою.
Запуск: python tests/create_dataset_template.py
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = os.path.join(os.path.dirname(__file__), "golden_dataset_template.xlsx")

# Кольори
COLOR_HEADER  = "1F4E79"   # темно-синій
COLOR_KB      = "DEEAF1"   # блакитний — KB
COLOR_COACH   = "E2EFDA"   # зелений — Коуч
COLOR_COMBO   = "FFF2CC"   # жовтий — Комбо
COLOR_FILL_H  = "FFFFFF"   # білий текст у заголовку

EXAMPLES = [
    # --- KB: HR та регламенти ---
    ("kb_001", "kb",    "HR / Звільнення",
     "Як оформити звільнення співробітника?",
     "Для звільнення потрібно оформити обхідний лист, підписати наказ про припинення трудового договору, провести розрахунок в бухгалтерії та видати трудову книжку.",
     "обходной лист, розрахунок, припинення"),

    ("kb_002", "kb",    "HR / Відпустка",
     "Як оформити щорічну відпустку?",
     "Необхідно подати заяву на ім'я керівника не пізніше ніж за 2 тижні. Відпустка оформлюється наказом. Тривалість — 24 календарні дні.",
     "заява, наказ, відпустка, 24 дні"),

    ("kb_003", "kb",    "HR / Лікарняний",
     "Що потрібно зробити якщо захворів?",
     "Потрібно повідомити керівника та HR в перший день хвороби. Після одужання надати лікарняний лист до бухгалтерії протягом 5 робочих днів.",
     "лікарняний, бухгалтерія, керівник"),

    ("kb_004", "kb",    "IT / Доступи",
     "Як отримати доступ до CRM?",
     "Заявку на доступ до CRM подає керівник відділу через IT-службу. Термін надання доступу — 1 робочий день.",
     "CRM, IT, доступ, заявка"),

    ("kb_005", "kb",    "Структура / Відділи",
     "Яка структура комерційного відділу?",
     "Комерційний відділ складається з регіональних менеджерів, key account менеджерів та відділу підтримки продажів.",
     "комерційний, менеджер, відділ"),

    # --- COACH: Препарати ---
    ("coach_001", "coach", "Препарат / Vitaran",
     "Розкажи про препарат Vitaran",
     "Vitaran — ін'єкційний препарат на основі PDRN (2%). Стимулює регенерацію через A2A-рецептори. Показання: постакне, рубці, суха шкіра, відновлення після процедур.",
     "PDRN, Vitaran, регенерація, постакне"),

    ("coach_002", "coach", "Препарат / Petaran",
     "Чим відрізняється Petaran від Vitaran?",
     "Petaran містить PDRN 1% + пептиди. Vitaran — чистий PDRN 2%. Petaran діє м'якше, підходить для чутливої шкіри та як підтримуюча терапія.",
     "Petaran, пептиди, чутлива шкіра"),

    ("coach_003", "coach", "Заперечення / Дорого",
     "Лікар каже що препарат дорогий, як відповісти?",
     "Порівняйте вартість курсу з результатом: один флакон Vitaran замінює 3-4 сесії з менш ефективним аналогом. Питайте: «Скільки коштує ваш час на ревізит пацієнта?»",
     "цінність, результат, аналог, вартість курсу"),

    ("coach_004", "coach", "Заперечення / Є аналог",
     "Лікар каже що є дешевший аналог",
     "Уточніть який саме аналог. Якщо PDRN — порівняйте концентрацію та виробника. Vitaran 2% PDRN від сертифікованого виробника — стабільний результат і мінімум побічних.",
     "аналог, концентрація, сертифікат, виробник"),

    ("coach_005", "coach", "Скрипт / Перший контакт",
     "Дай скрипт першого контакту з новим лікарем",
     "Представтесь, згадайте 1-2 клініки де вже працюєте. Задайте відкрите питання: «Які завдання зараз найактуальніші для ваших пацієнтів?» Слухайте — не презентуйте одразу.",
     "скрипт, перший контакт, відкрите питання"),

    # --- COMBO: Комбо-протоколи ---
    ("combo_001", "combo", "Комбо / Vitaran + біоревіталізація",
     "З чим комбінують Vitaran?",
     "Комбо-протокол: Vitaran + гіалуронова кислота (біоревіталізація). Послідовність: спочатку Vitaran для підготовки тканин, через 2 тижні — ГК. Ефект пролонгується на 30%.",
     "комбо, Vitaran, біоревіталізація, ГК"),

    ("combo_002", "combo", "Комбо / Постакне",
     "Які комбо-протоколи для постакне?",
     "Протокол постакне: Vitaran (3 сесії) + пілінг саліциловою кислотою. Або: Vitaran + PRP. Vitaran запускає регенерацію, пілінг/PRP прискорює оновлення клітин.",
     "постакне, Vitaran, PRP, пілінг, протокол"),

    ("combo_003", "combo", "Комбо / Petaran + ботулотоксин",
     "Чи можна поєднати Petaran з ботулотоксином?",
     "Так. Протокол: спочатку ботулотоксин (Neuronox/IUSE), через 14 днів — Petaran для зволоження та відновлення. Не вводити одночасно в одну зону.",
     "Petaran, ботулотоксин, Neuronox, поєднання"),
]

COLUMNS = [
    ("ID",              12),
    ("Режим",           10),
    ("Категорія",       22),
    ("Питання",         45),
    ("Очікувана відповідь (суть)",  55),
    ("Ключові слова для перевірки", 40),
]

MODE_COLORS = {"kb": COLOR_KB, "coach": COLOR_COACH, "combo": COLOR_COMBO}
MODE_LABELS = {"kb": "🔍 KB", "coach": "💼 Коуч", "combo": "🔗 Комбо"}


def thin_border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)


def build():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Golden Dataset"

    # Заголовок аркуша
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "EMET Bot — Golden Dataset для RAG-тестування"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Інструкція
    ws.merge_cells("A2:F2")
    inst = ws["A2"]
    inst.value = (
        "Заповни жовті комірки реальними даними. "
        "Кольори рядків: блакитний = KB (регламенти), зелений = Коуч (препарати), жовтий = Комбо-протоколи."
    )
    inst.font = Font(name="Calibri", italic=True, size=10, color="555555")
    inst.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 22

    # Заголовки колонок
    header_fill = PatternFill("solid", fgColor="2E75B6")
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[3].height = 24

    # Дані
    for row_idx, (id_, mode, category, question, answer, keywords) in enumerate(EXAMPLES, start=4):
        row_fill = PatternFill("solid", fgColor=MODE_COLORS[mode])
        values = [id_, MODE_LABELS[mode], category, question, answer, keywords]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.font = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border()
        ws.row_dimensions[row_idx].height = 55

    # Порожні рядки для заповнення
    for i in range(5):
        row_idx = len(EXAMPLES) + 4 + i
        mode_hint = ["kb", "kb", "coach", "combo", "combo"][i]
        row_fill = PatternFill("solid", fgColor=MODE_COLORS[mode_hint])
        placeholder = [
            f"new_{i+1:03d}", MODE_LABELS[mode_hint],
            "← категорія", "← введи своє питання",
            "← введи очікувану відповідь (суть, 1-3 речення)",
            "← ключові слова через кому"
        ]
        for col_idx, val in enumerate(placeholder, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.font = Font(name="Calibri", size=10, color="999999", italic=True)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border()
        ws.row_dimensions[row_idx].height = 55

    # Закріплюємо рядок заголовків
    ws.freeze_panes = "A4"

    wb.save(OUTPUT)
    print(f"✅ Шаблон збережено: {OUTPUT}")


if __name__ == "__main__":
    build()
